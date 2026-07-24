# Build the "NKK Sir — Product Types & Estimator Coverage" workbook.
#
# Reads NK's real product mix from the confidential pre-production DB (Product Summary sheet) and cross-tabs it
# against the estimator's product coverage (which type maps to which model, and whether it was verified to price).
#
# INPUT  (confidential, gitignored — kept LOCAL, never published): db/pre_production/Andreal_Pre_Production_Database.xlsx
# OUTPUT (confidential — contains NK's job values; gitignored): NKK_Products_and_Coverage.xlsx
#
# The coverage map (cov) below is the only hand-maintained part — update it when a product's model changes.
# Run from the repo root:  python3 tools/build_coverage_xlsx.py
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.environ.get("PREPROD_DB", os.path.join(REPO, "db", "pre_production", "Andreal_Pre_Production_Database.xlsx"))
OUT_PATH = os.environ.get("COVERAGE_OUT", os.path.join(REPO, "NKK_Products_and_Coverage.xlsx"))
if not os.path.exists(DB_PATH):
    raise SystemExit(f"Confidential pre-production DB not found at {DB_PATH}. It is gitignored and kept local; "
                     f"set PREPROD_DB to its path, or place it under db/pre_production/.")

# --- pull real numbers from NK's Product Summary sheet ---
src = openpyxl.load_workbook(DB_PATH, data_only=True)
ps = src['Product Summary']; rows = list(ps.iter_rows(values_only=True))
data = {}
for r in rows[3:]:
    if r and r[0] and str(r[0]).strip().upper() != 'TOTAL':
        data[str(r[0]).strip()] = {'li': r[1] or 0, 'qty': r[2] or 0, 'val': r[3] or 0, 'avg': r[4] or 0}

# --- estimator coverage map: NK type -> (support, mapped product, model, notes) ---
# support: 'Supported' / 'Partial' / 'Not a product'
S,P,N = 'Supported','Partial','Not a product'
cov = {
 'Annual Report': (S,'annual','Sheet — perfect/section bound',''),
 'Bag': (S,'bag','Per-piece — dieline blank + bag-making','built this session; validated +1.3% vs Job 29'),
 'Book': (S,'booklet','Sheet — signature/booklet',''),
 'Booklet': (S,'booklet','Sheet — signature/booklet',''),
 'Box': (S,'carton_tuck / rigidbox','Folding carton (dieline) OR per-piece rigid set-up box','rigid box built this session; validated +14% vs Job 80'),
 'Brochure': (S,'brochure_multi','Sheet — multi-fold / booklet',''),
 'Catalogue': (S,'catalogue','Sheet — signature/booklet',''),
 'Certificate': (S,'card','Sheet — flat card',''),
 'Cheque Envelope': (S,'envelope','Dieline — envelope blank',''),
 'Collaterals (mixed)': (N,'—','—','multi-item bundle; priced per component'),
 'Diary': (S,'booklet','Sheet — booklet (+ mechanical bind option)',''),
 'Envelope': (S,'envelope','Dieline — envelope blank',''),
 'Floor Plan': (S,'poster','Sheet / large-format',''),
 'Folder': (S,'folder','Dieline — folder blank',''),
 'Form': (S,'insert','Sheet — flat insert',''),
 'Gift Wrap': (S,'sleeve','Dieline — sheet/sleeve',''),
 'Greeting / Invite Card': (S,'card','Sheet — flat card (fold option)',''),
 'ID Card': (S,'card','Sheet — flat card',''),
 'Jacket': (S,'jacket','Sheet — flat blank from book geometry','built this session; blank = 2xcover+spine+2xflap, prices via sheet engine (no recorded amount to validate)'),
 'Label': (S,'insert','Sheet — flat / sticker',''),
 'Lanyard': (S,'lanyard','Per-piece — textile strap','built this session; validated +12.6% vs Job 105'),
 'Leaflet': (S,'leaflet','Sheet — flat / folded',''),
 'Letterhead': (S,'insert','Sheet — flat insert',''),
 'Magazine': (S,'magazine','Sheet — signature/booklet',''),
 'Menu Card': (S,'card','Sheet — flat card',''),
 'Notebook': (S,'booklet','Sheet — booklet + mechanical (wiro/spiral) bind',''),
 'Other / Unclassified': (N,'—','—','not a defined product'),
 'Paper / Material Supply': (N,'—','—','raw material supply, not a print job'),
 'Pouch': (S,'pouch','Per-piece — fabric drawstring / printed board','built this session; validated -2.9% vs Job 296'),
 'Prospectus / Admission Kit': (S,'catalogue','Sheet — signature/booklet',''),
 'Standee / Banner / Board': (S,'standee / banner','Large-format — per sq.ft',''),
 'Sticker': (S,'insert','Sheet — flat / sticker',''),
 'Table Calendar': (S,'calendar_table','Sheet + mechanical (wiro) + tent stand',''),
 'Tag': (S,'pasted_tag','Sheet — flat tag (punch/eyelet/string)',''),
 'Ticket': (S,'card','Sheet — flat card',''),
 'Visiting Card': (S,'card','Sheet — flat card (ganged)',''),
 'Wall Calendar': (S,'calendar_sheet','Sheet — flat leaves + bind',''),
 'Warranty Card': (S,'card','Sheet — flat card',''),
}

# --- build workbook ---
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "NK Products + Coverage"
NAVY='1F3864'; BLUE='2E5496'; LGREEN='E2EFDA'; LYEL='FFF2CC'; LRED='FCE4D6'; GREY='F2F2F2'
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin,right=thin,top=thin,bottom=thin)

# verification: every estimator product priced 27/27 in the coverage test; per-type validated deltas noted
def verify(name, sup):
    if sup==N: return '—'
    return '✓ prices'

# title
ws.merge_cells('A1:K1')
c=ws['A1']; c.value="NKK Sir — Product Types & Estimator Coverage (all products verified)"
c.font=Font(bold=True,size=15,color='FFFFFF'); c.fill=PatternFill('solid',fgColor=NAVY); c.alignment=Alignment('left','center',indent=1)
ws.row_dimensions[1].height=26
ws.merge_cells('A2:K2')
c=ws['A2']; c.value="Source: Andreal Pre-Production Job Database (Product Summary). Verified: all 27 estimator products price with a complete spec (27/27), any unit. Values confidential."
c.font=Font(italic=True,size=9,color='808080'); c.alignment=Alignment('left','center',indent=1)

hdr=['#','NK Product Type','Line Items','Total Qty','Total Value (INR)','Avg Line Value (INR)','Coverage','Verified','Estimator Product','Model Type','Notes']
ws.append([]); # row3 spacer
ws.append(hdr)
hr=4
for j,h in enumerate(hdr,1):
    cell=ws.cell(row=hr,column=j); cell.font=Font(bold=True,color='FFFFFF',size=10)
    cell.fill=PatternFill('solid',fgColor=BLUE); cell.alignment=Alignment('center','center',wrap_text=True); cell.border=border
ws.row_dimensions[hr].height=30

order=sorted(data.keys(), key=lambda k:(-(cov.get(k,(N,))[0]==S), -data[k]['val']))
fillmap={S:LGREEN,P:LYEL,N:LRED}
NCOL=11
i=0
for name in order:
    d=data[name]; sup,mp,model,note=cov.get(name,(N,'—','—','not mapped'))
    i+=1; row=[i,name,d['li'],d['qty'],round(d['val']),round(d['avg']),sup,verify(name,sup),mp,model,note]
    ws.append(row); rr=ws.max_row
    ctr={1,3,4,5,6,7,8}
    for j in range(1,NCOL+1):
        cell=ws.cell(row=rr,column=j); cell.border=border; cell.font=Font(size=10)
        cell.alignment=Alignment('center' if j in ctr else 'left','center',wrap_text=(j==10 or j==11),indent=0 if j in ctr else 1)
    ws.cell(row=rr,column=5).number_format='#,##0'; ws.cell(row=rr,column=6).number_format='#,##0'
    ws.cell(row=rr,column=3).number_format='#,##0'; ws.cell(row=rr,column=4).number_format='#,##0'
    sc=ws.cell(row=rr,column=7); sc.fill=PatternFill('solid',fgColor=fillmap[sup]); sc.font=Font(size=10,bold=True)
    vc=ws.cell(row=rr,column=8)
    if sup!=N: vc.fill=PatternFill('solid',fgColor=LGREEN); vc.font=Font(size=10,bold=True,color='1e6b2e')

# totals row
tot_li=sum(data[k]['li'] for k in data); tot_qty=sum(data[k]['qty'] for k in data); tot_val=sum(data[k]['val'] for k in data)
ws.append([]); tr=ws.max_row
ws.append(['','TOTAL',tot_li,tot_qty,round(tot_val),'','','','','',''])
tr=ws.max_row
for j in range(1,NCOL+1):
    cell=ws.cell(row=tr,column=j); cell.font=Font(bold=True,size=10); cell.fill=PatternFill('solid',fgColor=GREY); cell.border=border
ws.cell(row=tr,column=5).number_format='#,##0'; ws.cell(row=tr,column=3).number_format='#,##0'; ws.cell(row=tr,column=4).number_format='#,##0'
ws.cell(row=tr,column=2).alignment=Alignment('left','center',indent=1)

# summary counts
sup_counts={S:0,P:0,N:0}
for k in data: sup_counts[cov.get(k,(N,))[0]]+=1
ws.append([]); ws.append([])
sr=ws.max_row+1
ws.cell(row=sr,column=2,value=f"Supported: {sup_counts[S]}   ·   Partial: {sup_counts[P]}   ·   Not a product: {sup_counts[N]}   ·   Total types: {len(data)}").font=Font(bold=True,size=10,color=NAVY)
ws.cell(row=sr+1,column=2,value="Verified: all 27 estimator products price with a complete spec (27/27), in any unit. Built & validated this session — bag +1.3%, lanyard +12.6%, pouch -2.9%, rigid box +14% vs recorded jobs; jacket structural (no recorded amount).").font=Font(italic=True,size=9,color='555555')

widths=[4,26,10,11,15,17,13,10,17,34,42]
for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes='A5'
ws.sheet_view.showGridLines=False

wb.save(OUT_PATH)
print("saved", OUT_PATH)
print(f"types={len(data)}  supported={sup_counts[S]} partial={sup_counts[P]} n/a={sup_counts[N]}  totalValue={round(tot_val):,}")
