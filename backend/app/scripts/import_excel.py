"""Import the Rate_Masters.xlsx workbook into PostgreSQL.

Header-driven, so it tolerates column reordering and the banner/subtitle/footnote
rows the workbook wraps every sheet in. Vendor *names* in the sheets are
auto-upserted into vendors() and linked by FK, so relationships are mapped
correctly without manual data entry.

Targets the staging tables in db/rate_masters.sql:
    Lamination/UV/Foiling/Embossing/Die Cutting/Binding -> postpress_rates
    Plate                                                -> plate_rates
    Packaging                                            -> packaging_rates
    Wastage / Margin / GST Rules                         -> *_rules_master

Usage:
    python -m app.scripts.import_excel --file "C:/Users/AR04/Downloads/Rate_Masters.xlsx"
    python -m app.scripts.import_excel --file Rate_Masters.xlsx --dry-run
    python -m app.scripts.import_excel --file Rate_Masters.xlsx --database-url postgresql+psycopg://...

--dry-run parses and prints a per-sheet summary without touching the database
(useful to validate parsing even before Postgres is wired up).
"""
from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal, InvalidOperation

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")

from sqlalchemy import create_engine, text

# --------------------------------------------------------------------------
# Per-sheet configuration.
#   target       : staging table name
#   category     : fixed discriminator value for postpress_rates (else None)
#   vendor_field : header alias that carries the vendor name (forward-filled);
#                  None for the rules sheets
#   columns      : {normalized_header: (db_column, kind)}
#                  kind in {text, decimal, int, bool} controls coercion
# --------------------------------------------------------------------------
_POSTPRESS_COLUMNS = {
    "vendor": ("vendor_name", "text"),
    "operation type": ("operation_type", "text"),
    "customer job ref": ("job_ref", "text"),
    "pages": ("pages", "text"),
    "qty": ("qty", "decimal"),
    "size in": ("size_text", "text"),
    "pricing unit": ("pricing_unit", "text"),
    "rate pcs": ("rate_per_pcs", "decimal"),
    "rate sheet": ("rate_per_sheet", "decimal"),
    "rate sq in": ("rate_per_sqin", "decimal"),
    "min charge": ("min_charge", "decimal"),
    "qty break min": ("qty_break_min", "decimal"),
    "qty break rate": ("qty_break_rate", "decimal"),
    "active": ("active", "bool"),
}

SHEET_CONFIG: dict[str, dict] = {
    "Lamination Rate Master": {"target": "postpress_rates", "category": "lamination",
                               "vendor_field": "vendor", "columns": _POSTPRESS_COLUMNS},
    "UV Rate Master":         {"target": "postpress_rates", "category": "uv",
                               "vendor_field": "vendor", "columns": _POSTPRESS_COLUMNS},
    "Foiling Rate Master":    {"target": "postpress_rates", "category": "foiling",
                               "vendor_field": "vendor", "columns": _POSTPRESS_COLUMNS},
    "Embossing Rate Master":  {"target": "postpress_rates", "category": "embossing",
                               "vendor_field": "vendor", "columns": _POSTPRESS_COLUMNS},
    "Die Cutting Rate Master":{"target": "postpress_rates", "category": "die_cutting",
                               "vendor_field": "vendor", "columns": _POSTPRESS_COLUMNS},
    "Binding Rate Master":    {"target": "postpress_rates", "category": "binding",
                               "vendor_field": "vendor", "columns": _POSTPRESS_COLUMNS},
    "Plate Rate Master": {
        "target": "plate_rates", "category": None, "vendor_field": "vendor",
        "columns": {
            "vendor": ("vendor_name", "text"),
            "machine": ("machine", "text"),
            "machine type": ("machine_type", "text"),
            "max sheet size": ("max_sheet_size", "text"),
            "colours": ("colours", "text"),
            "plate charge colour small machine": ("plate_charge_small", "decimal"),
            "plate charge colour big machine": ("plate_charge_big", "decimal"),
        },
    },
    "Packaging Rate Master": {
        "target": "packaging_rates", "category": None, "vendor_field": "vendor",
        "columns": {
            "vendor": ("vendor_name", "text"),
            "category": ("category", "text"),
            "material type": ("material_type", "text"),
            "description": ("description", "text"),
            "size l in": ("size_l_in", "decimal"),
            "size w in": ("size_w_in", "decimal"),
            "unit": ("unit", "text"),
            "rate unit": ("rate_per_unit", "decimal"),
            "moq": ("moq", "decimal"),
            "lead time days": ("lead_time_days", "int"),
        },
    },
    "Wastage Rules": {
        "target": "wastage_rules_master", "category": None, "vendor_field": None,
        "columns": {
            "process": ("process", "text"),
            "wastage basis": ("wastage_basis", "text"),
            "wastage sheets": ("wastage_value", "text"),
            "notes": ("notes", "text"),
        },
    },
    "Margin Rules": {
        "target": "margin_rules_master", "category": None, "vendor_field": None,
        "columns": {
            "cost component job tier": ("cost_component", "text"),
            "margin basis": ("margin_basis", "text"),
            "margin markup on cost": ("margin_value", "text"),
            "notes": ("notes", "text"),
        },
    },
    "GST Rules": {
        "target": "gst_rules_master", "category": None, "vendor_field": None,
        "columns": {
            "item service": ("item_service", "text"),
            "indicative hsn sac": ("hsn_sac", "text"),
            "indicative gst": ("gst_value", "text"),
            "notes": ("notes", "text"),
        },
    },
}


# --------------------------------------------------------------------------
# Parsing helpers
# --------------------------------------------------------------------------
def norm(s) -> str:
    """Normalize a header/label: lower-case, drop symbols & units, collapse spaces."""
    if s is None:
        return ""
    s = str(s).replace("₹", " ")
    s = re.sub(r"[^a-z0-9]+", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def to_decimal(v):
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace("₹", "").replace(",", "").replace("%", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None  # free-text like "1500/1000" — kept only where column is text


def to_int(v):
    d = to_decimal(v)
    return int(d) if d is not None else None


def to_bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in {"yes", "y", "true", "1", "active"}:
        return True
    if s in {"no", "n", "false", "0", "inactive"}:
        return False
    return None


_COERCE = {"text": to_text, "decimal": to_decimal, "int": to_int, "bool": to_bool}


def find_header_row(rows: list[tuple]) -> int:
    """First row with >= 3 non-empty string cells (skips merged banner/subtitle)."""
    for i, row in enumerate(rows):
        strings = sum(1 for c in row if isinstance(c, str) and c.strip())
        if strings >= 3:
            return i
    return -1


def parse_sheet(ws, cfg: dict) -> list[dict]:
    """Return a list of {db_column: value} dicts for the sheet's data rows."""
    rows = list(ws.iter_rows(values_only=True))
    h = find_header_row(rows)
    if h < 0:
        return []

    header = rows[h]
    # map db_column -> source column index, via normalized header match
    col_index: dict[str, tuple[int, str]] = {}
    vendor_db_col = None
    for idx, cell in enumerate(header):
        key = norm(cell)
        if key in cfg["columns"]:
            db_col, kind = cfg["columns"][key]
            col_index[db_col] = (idx, kind)
            if cfg["vendor_field"] and key == cfg["vendor_field"]:
                vendor_db_col = db_col

    records: list[dict] = []
    last_vendor = None
    for r_offset, row in enumerate(rows[h + 1:], start=h + 2):  # 1-based sheet row no.
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty < 2:          # blank row or single-cell footnote
            continue

        rec: dict = {}
        for db_col, (idx, kind) in col_index.items():
            raw = row[idx] if idx < len(row) else None
            rec[db_col] = _COERCE[kind](raw)

        # forward-fill vendor across merged/continuation rows
        if vendor_db_col:
            if rec.get(vendor_db_col):
                last_vendor = rec[vendor_db_col]
            else:
                rec[vendor_db_col] = last_vendor

        if cfg["category"]:
            rec["operation_category"] = cfg["category"]
        rec["source_row"] = r_offset
        records.append(rec)
    return records


# --------------------------------------------------------------------------
# Vendor resolution
# --------------------------------------------------------------------------
# Variant spellings found in the master sheets that mean an existing canonical
# vendor in Vendor_Master. Keys are _key()-normalised (UPPER, single-spaced).
# Without these the importer would create duplicate VEN### rows.
VENDOR_ALIASES: dict[str, str] = {
    "KRISHNA VANIJYA PVT. LTD.": "VEN003",   # vs "KRISHNA VANIJYA PVT LTD."
    "F.AHMED PAPER HOUSE": "VEN006",         # vs "F. AHMED & CO."
    "A.K. TRADING PVT LTD": "VEN009",        # vs "A.K. TRADING PVT LTD." (trailing .)
    "COSMOS CARD AND BOX": "VEN023",         # vs "COSMOS CARD BOARD BOX"
    "PRAKSH CHOWDHURY": "VEN028",            # vs "PRAKASH CHOWDHURY"
    "S.B. TRADE": "VEN040",                  # vs "S.B TRADE"
}


class VendorResolver:
    """Upserts vendor names into vendors() and hands back vendor_ids."""

    def __init__(self, conn):
        self.conn = conn
        self.by_name: dict[str, str] = {}
        self._next = 1
        if conn is not None:
            for vid, vname in conn.execute(text("SELECT vendor_id, vendor_name FROM vendors")):
                if vname:
                    self.by_name[self._key(vname)] = vid
                m = re.fullmatch(r"VEN(\d+)", vid or "")
                if m:
                    self._next = max(self._next, int(m.group(1)) + 1)

    @staticmethod
    def _key(name: str) -> str:
        return re.sub(r"\s+", " ", str(name).strip().upper())

    def resolve(self, name: str | None) -> str | None:
        if not name:
            return None
        k = self._key(name)
        if k in VENDOR_ALIASES:               # known variant → canonical id
            return VENDOR_ALIASES[k]
        if k in self.by_name:
            return self.by_name[k]
        vid = f"VEN{self._next:03d}"
        self._next += 1
        self.by_name[k] = vid
        if self.conn is not None:
            self.conn.execute(
                text("INSERT INTO vendors (vendor_id, vendor_name, vendor_type, active) "
                     "VALUES (:id, :name, 'imported', TRUE)"),
                {"id": vid, "name": str(name).strip()},
            )
        return vid


# --------------------------------------------------------------------------
# Driver
# --------------------------------------------------------------------------
def run(file_path: str, database_url: str | None, dry_run: bool) -> int:
    wb = load_workbook(file_path, data_only=True, read_only=True)

    engine = conn = trans = resolver = None
    if not dry_run:
        if not database_url:
            from app.config import settings
            database_url = settings.database_url
        engine = create_engine(database_url, future=True)
        conn = engine.connect()
        trans = conn.begin()
        resolver = VendorResolver(conn)
    else:
        resolver = VendorResolver(None)

    grand_total = 0
    new_vendors_before = len(resolver.by_name)
    try:
        for sheet_name, cfg in SHEET_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                print(f"  ! sheet not found, skipping: {sheet_name}")
                continue
            ws = wb[sheet_name]
            records = parse_sheet(ws, cfg)

            for rec in records:
                if "vendor_name" in rec:
                    rec["vendor_id"] = resolver.resolve(rec.get("vendor_name"))
                rec["source_sheet"] = sheet_name

            if not dry_run:
                conn.execute(text(f"DELETE FROM {cfg['target']}"))  # full reload
                if records:
                    cols = sorted({c for r in records for c in r})
                    stmt = text(
                        f"INSERT INTO {cfg['target']} ({', '.join(cols)}) "
                        f"VALUES ({', '.join(':' + c for c in cols)})"
                    )
                    conn.execute(stmt, [{c: r.get(c) for c in cols} for r in records])

            grand_total += len(records)
            print(f"  {sheet_name:28s} -> {cfg['target']:22s} {len(records):4d} rows")

        if not dry_run:
            trans.commit()
        new_vendors = len(resolver.by_name) - new_vendors_before
        mode = "DRY-RUN (no writes)" if dry_run else "committed"
        print(f"\nDone [{mode}]: {grand_total} rows across {len(SHEET_CONFIG)} sheets, "
              f"{new_vendors} new vendor(s) resolved.")
        return 0
    except Exception:
        if trans is not None:
            trans.rollback()
        raise
    finally:
        if conn is not None:
            conn.close()


def main() -> None:
    ap = argparse.ArgumentParser(description="Import Rate_Masters.xlsx into PostgreSQL.")
    ap.add_argument("--file", required=True, help="Path to Rate_Masters.xlsx")
    ap.add_argument("--database-url", default=None, help="Override DATABASE_URL")
    ap.add_argument("--dry-run", action="store_true", help="Parse & summarise, no DB writes")
    args = ap.parse_args()
    sys.exit(run(args.file, args.database_url, args.dry_run))


if __name__ == "__main__":
    main()
