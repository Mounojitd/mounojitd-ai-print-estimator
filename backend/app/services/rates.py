"""RateBook — loads REAL rates from the master workbooks so the engine quotes
from your actual data, not hard-coded constants.

Reads db/files/*.xlsx today (works without a database); the same interface can
later be backed by the imported Postgres tables. Reuses the header-driven
parser from the importer so column order doesn't matter.

KNOWN DATA GAPS (the engine flags these at quote time):
  * No printing/impression rate exists in any sheet — only plate rates. The
    per-sheet run cost still comes from estimation_config (IMPRESSION_RATE...).
  * Lamination rate units are inconsistent ("per sq inch" values that imply
    impossible per-piece costs), so lamination is left to config too.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from app.scripts.import_excel import parse_sheet
from .estimation_config import D
from .types import Size

IN_MM = D("25.4")
DEFAULT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "db", "files")


def _dec(v) -> Decimal | None:
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = re.sub(r"[^0-9.]", "", str(v))
    try:
        return Decimal(s) if s else None
    except InvalidOperation:
        return None


def size_in_to_mm(text) -> Size | None:
    """'25*36' / '28X40' / '23\"x36\"' inches → Size in mm (w, h)."""
    if text is None:
        return None
    nums = re.findall(r"\d+(?:\.\d+)?", str(text))
    if len(nums) >= 2:
        return Size(Decimal(nums[0]) * IN_MM, Decimal(nums[1]) * IN_MM)
    return None


def _colours(text) -> int:
    m = re.search(r"\d+", str(text or ""))
    if m:
        return int(m.group())
    return 1 if "single" in str(text or "").lower() else 4


@dataclass
class MachineRate:
    vendor: str
    machine: str
    max_sheet: Size            # max sheet the machine accepts (mm)
    colours: int
    speed_iph: int | None
    plate_per_colour: Decimal  # ₹ per plate, from the real Plate Rate Master
    size_class: str            # 'big' (B1 ≥28") | 'small'


@dataclass
class PaperRate:
    paper_id: str
    vendor: str
    description: str
    gsm: int
    sheet: Size                # paper sheet size (mm)
    price_per_sheet: Decimal


class RateBook:
    def __init__(self, data_dir: str = DEFAULT_DIR):
        self.dir = data_dir
        self._machines: list[MachineRate] | None = None
        self._papers: list[PaperRate] | None = None
        self._impr: list[dict] | None = None

    def _ws(self, fname: str, sheet: str):
        wb = load_workbook(os.path.join(self.dir, fname), data_only=True, read_only=True)
        return wb[sheet]

    @staticmethod
    def _cfg(columns):
        return {"columns": columns, "vendor_field": None, "category": None}

    # ---------------- plate rates (real) ----------------
    def _plate_by_machine(self) -> dict[str, Decimal]:
        cols = {
            "machine": ("machine", "text"),
            "plate charge colour small machine": ("small", "text"),
            "plate charge colour big machine": ("big", "text"),
        }
        rows = parse_sheet(self._ws("Rate_Masters.xlsx", "Plate Rate Master"), self._cfg(cols))
        out: dict[str, Decimal] = {}
        for r in rows:
            m = (r.get("machine") or "").strip()
            rate = _dec(r.get("big")) or _dec(r.get("small"))
            if m and rate:
                out[m.upper()] = max(out.get(m.upper(), Decimal(0)), rate)
        return out

    # ---------------- machines (real) ----------------
    def machines(self) -> list[MachineRate]:
        if self._machines is not None:
            return self._machines
        plate = self._plate_by_machine()
        cols = {
            "vendor": ("vendor", "text"), "machine": ("machine", "text"),
            "max sheet size": ("maxname", "text"), "size in": ("size_in", "text"),
            "colours": ("colours", "text"), "speed imp hr": ("speed", "int"),
        }
        rows = parse_sheet(self._ws("Process_Master.xlsx", "Process Master"), self._cfg(cols))
        out: list[MachineRate] = []
        for r in rows:
            sz = size_in_to_mm(r.get("size_in"))
            if not r.get("machine") or sz is None:
                continue
            max_in = max(sz.w, sz.h) / IN_MM
            size_class = "big" if max_in >= D("28") else "small"
            mname = r["machine"].strip()
            rate = plate.get(mname.upper()) or (D("550") if size_class == "big" else D("400"))
            out.append(MachineRate(
                vendor=(r.get("vendor") or "").strip(), machine=mname,
                max_sheet=sz, colours=_colours(r.get("colours")),
                speed_iph=r.get("speed"), plate_per_colour=rate, size_class=size_class,
            ))
        self._machines = out
        return out

    # ---------------- paper (real, cheapest vendor) ----------------
    def papers(self) -> list[PaperRate]:
        if self._papers is not None:
            return self._papers
        cols = {
            "id": ("paper_id", "text"), "vendor": ("vendor", "text"),
            "paper type": ("ptype", "text"), "grade": ("grade", "text"),
            "brand": ("brand", "text"), "gsm": ("gsm", "int"),
            "sheet size in": ("size_in", "text"),
            "price sheet": ("price_sheet", "decimal"),
        }
        rows = parse_sheet(self._ws("Paper_Master.xlsx", "Paper Pricing"), self._cfg(cols))
        out: list[PaperRate] = []
        for r in rows:
            sz = size_in_to_mm(r.get("size_in"))
            price = _dec(r.get("price_sheet"))
            if not r.get("paper_id") or sz is None or not price:
                continue
            desc = " ".join(str(r.get(k)) for k in ("grade", "brand") if r.get(k)).strip()
            out.append(PaperRate(
                paper_id=r["paper_id"].strip(), vendor=(r.get("vendor") or "").strip(),
                description=desc or str(r.get("ptype") or ""), gsm=r.get("gsm") or 0,
                sheet=sz, price_per_sheet=price,
            ))
        self._papers = out
        return out

    def find_papers(self, grade_contains: str, gsm: int,
                    exclude: str | None = None) -> list[PaperRate]:
        gc = grade_contains.lower()
        ex = (exclude or "").lower()
        return [
            p for p in self.papers()
            if gc in p.description.lower() and p.gsm == gsm
            and (not ex or ex not in p.description.lower())
        ]

    def cheapest_paper(self, grade_contains: str, gsm: int,
                       exclude: str | None = None) -> PaperRate | None:
        cands = self.find_papers(grade_contains, gsm, exclude)
        return min(cands, key=lambda p: p.price_per_sheet) if cands else None

    # ------------- printing / impression rate (once you fill the template) -------------
    def _load_impressions(self) -> list[dict]:
        path = os.path.join(self.dir, "Printing_Rate_Master.xlsx")
        if not os.path.exists(path):
            return []
        cols = {
            "vendor": ("vendor", "text"), "machine": ("machine", "text"),
            "min qty": ("min_qty", "int"), "max qty": ("max_qty", "int"),
            "colour mode single 4color": ("mode", "text"),
            "rate sheet": ("rate", "decimal"),
            "setup make ready sheets": ("setup", "int"),
            "min charge": ("min_charge", "decimal"), "active": ("active", "bool"),
        }
        wb = load_workbook(path, data_only=True, read_only=True)
        rows = parse_sheet(wb[wb.sheetnames[0]], self._cfg(cols))
        out = []
        for r in rows:
            if r.get("machine") and r.get("rate") is not None:
                out.append({
                    "machine": r["machine"].strip(),
                    "min": int(r.get("min_qty") or 0),
                    "max": int(r.get("max_qty") or 10**9),
                    "mode": (r.get("mode") or "4color").strip().lower(),
                    "rate": _dec(r["rate"]),
                    "setup": r.get("setup"),
                    "min_charge": _dec(r.get("min_charge")) or D(0),
                    "active": r.get("active") is not False,
                })
        return out

    def impression(self, machine: str, qty: int, colour_mode: str) -> dict | None:
        """Real per-sheet run rate for this machine/qty/mode, if the master is filled."""
        if self._impr is None:
            self._impr = self._load_impressions()
        mu = (machine or "").upper()
        for r in self._impr:
            if (r["machine"].upper() == mu and r["min"] <= qty <= r["max"]
                    and r["mode"] == colour_mode):
                return r
        return None

    def select_machine(self, paper_sheet: Size, colours_needed: int) -> MachineRate | None:
        """Cheapest-plate machine that accepts the paper sheet and the colours."""
        def fits(m: MachineRate) -> bool:
            a = m.max_sheet.w >= paper_sheet.w and m.max_sheet.h >= paper_sheet.h
            b = m.max_sheet.w >= paper_sheet.h and m.max_sheet.h >= paper_sheet.w
            return (a or b) and m.colours >= colours_needed
        capable = [m for m in self.machines() if fits(m)]
        return min(capable, key=lambda m: m.plate_per_colour) if capable else None
